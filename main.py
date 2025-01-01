from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout, QHBoxLayout, QPushButton,
    QLineEdit, QLabel, QComboBox, QListWidget, QWidget, QFileDialog, QMessageBox, QFormLayout, QGridLayout
)
from PyQt6.QtCore import Qt
import sys
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

class ExcelAddressTool(QMainWindow):

    def __init__(self):
        super().__init__()
        self.selected_order_id = None  # 선택된 항목의 ID 저장
        self.init_ui()

    def init_ui(self):
        super().__init__()

        self.setWindowTitle("Excel Address Tool")
        self.setGeometry(100, 100, 1200, 800)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        self.layout = QVBoxLayout(self.central_widget)

        # Database Initialization
        self.init_database()

        # Top Panel: Excel Table Display
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(14)
        self.table_widget.setHorizontalHeaderLabels([
            "id", "보내는분 성명", "보내는분 전화번호", "보내는분 주소(전체, 분할)", "보내는분 기타연락처",
            "받는분 성명", "받는분 전화번호", "받는분 주소(전체, 분할)", "받는분 기타연락처",
            "품목명", "갯수", "기본운임", "배송메세지1", "운임구분"
        ])
        self.layout.addWidget(self.table_widget, stretch=2)
        self.table_widget.cellClicked.connect(self.table_item_clicked)
        # Load Data from Database into Table Widget
        self.load_data()

        # Middle Panel: Input Form and Address List
        self.middle_layout = QHBoxLayout()

        # Input Form
        self.input_form_layout = QGridLayout()

        self.sender_name_input = QLineEdit()
        self.sender_phone_input = QLineEdit()
        self.sender_address_input = QLineEdit()
        self.receiver_name_input = QLineEdit()
        self.receiver_phone_input = QLineEdit()
        self.receiver_address_input = QLineEdit()
        self.item_name_input = QComboBox()  # 품목명 입력창을 QComboBox로 변경
        self.item_name_input.setEditable(True)  # 텍스트 입력을 허용
        self.quantity_input = QLineEdit()

        # Default values
        self.sender_name_input.setText("행복한감귤농장")
        self.sender_phone_input.setText("010-5772-4798")
        self.sender_address_input.setText("제주도 서귀포시 남원읍 태위로 603번길 18-5")

        self.sender_name_search_button = QPushButton("검색")
        self.receiver_name_search_button = QPushButton("검색")
        self.edit_button = QPushButton("수정")
        self.cancel_button = QPushButton("취소")
        self.delete_button = QPushButton("삭제")
        self.add_button = QPushButton("추가")

        self.edit_button.hide()
        self.cancel_button.hide()
        self.delete_button.hide()

        self.input_form_layout.addWidget(QLabel("보내는분 성명:"), 0, 0)
        self.input_form_layout.addWidget(self.sender_name_input, 0, 1, 1, 2)
        self.input_form_layout.addWidget(self.sender_name_search_button, 0, 3)

        self.input_form_layout.addWidget(QLabel("보내는분 전화번호:"), 1, 0)
        self.input_form_layout.addWidget(self.sender_phone_input, 1, 1, 1, 3)

        self.input_form_layout.addWidget(QLabel("보내는분 주소(전체, 분할):"), 2, 0)
        self.input_form_layout.addWidget(self.sender_address_input, 2, 1, 1, 3)

        self.input_form_layout.addWidget(QLabel("받는분 성명:"), 3, 0)
        self.input_form_layout.addWidget(self.receiver_name_input, 3, 1, 1, 2)
        self.input_form_layout.addWidget(self.receiver_name_search_button, 3, 3)

        self.input_form_layout.addWidget(QLabel("받는분 전화번호:"), 4, 0)
        self.input_form_layout.addWidget(self.receiver_phone_input, 4, 1, 1, 3)

        self.input_form_layout.addWidget(QLabel("받는분 주소(전체, 분할):"), 5, 0)
        self.input_form_layout.addWidget(self.receiver_address_input, 5, 1, 1, 3)

        self.input_form_layout.addWidget(QLabel("품목명:"), 6, 0)
        self.input_form_layout.addWidget(self.item_name_input, 6, 1, 1, 3)

        # 품목명을 드롭다운으로 추가
        self.item_name_input.addItems([
            "조생귤 10kg", "조생귤 5kg", "저농약귤 10kg", "저농약귤 5kg", "한라봉 10kg", "한라봉 5kg", 
            "황금향 10kg", "황금향 5kg", "청견 10kg", "청견 5kg", "못난이 귤 10kg", "대과 귤 10kg"
        ])

        self.input_form_layout.addWidget(QLabel("갯수:"), 7, 0)
        self.input_form_layout.addWidget(self.quantity_input, 7, 1, 1, 3)

        # 버튼 배치를 같은 행에 정렬
        self.input_form_layout.addWidget(self.add_button, 8, 3, 1, 1, alignment=Qt.AlignmentFlag.AlignRight)
        self.input_form_layout.addWidget(self.edit_button, 8, 3, 1, 1, alignment=Qt.AlignmentFlag.AlignRight)
        self.input_form_layout.addWidget(self.cancel_button, 8, 2, 1, 1, alignment=Qt.AlignmentFlag.AlignRight)
        self.input_form_layout.addWidget(self.delete_button, 8, 1, 1, 1, alignment=Qt.AlignmentFlag.AlignRight)


        self.input_panel = QWidget()
        self.input_panel.setLayout(self.input_form_layout)

        # Address List
        self.address_list_layout = QVBoxLayout()
        self.address_list_label = QLabel("이름으로 찾은 주소 목록:")
        self.address_list_widget = QListWidget()
        # 너비를 원하는 만큼 조정 (예: 400px로 설정)
        self.address_list_widget.setFixedWidth(500)  # 원하는 너비로 수정
        self.address_list_widget.itemClicked.connect(self.fill_address_and_phone_field)  # Connect item click event
        self.address_list_layout.addWidget(self.address_list_label)
        self.address_list_layout.addWidget(self.address_list_widget)

        self.address_panel = QWidget()
        self.address_panel.setLayout(self.address_list_layout)

        # Add to Middle Layout
        self.middle_layout.addWidget(self.input_panel, stretch=3)
        self.middle_layout.addWidget(self.address_panel, stretch=1)

        self.layout.addLayout(self.middle_layout, stretch=3)

        # Bottom Panel: Buttons
        self.bottom_panel = QHBoxLayout()
        self.reset_button = QPushButton("초기화")
        self.export_button = QPushButton("엑셀 추출")
        self.bottom_panel.addWidget(self.reset_button)
        self.bottom_panel.addWidget(self.export_button)
        self.layout.addLayout(self.bottom_panel)

        # Button Connections
        self.reset_button.clicked.connect(self.confirm_reset)
        self.export_button.clicked.connect(self.export_to_excel)
        self.add_button.clicked.connect(self.add_entry)
        self.edit_button.clicked.connect(self.edit_entry)
        self.cancel_button.clicked.connect(self.cancel_edit)
        self.delete_button.clicked.connect(self.confirm_delete)
        self.sender_name_search_button.clicked.connect(lambda: self.search_address(self.sender_name_input.text()))
        self.receiver_name_search_button.clicked.connect(lambda: self.search_address(self.receiver_name_input.text()))

        # 엔터 키를 눌렀을 때 검색이 자동으로 되도록 연결
        self.sender_name_input.returnPressed.connect(lambda: self.search_address(self.sender_name_input.text()))
        self.receiver_name_input.returnPressed.connect(lambda: self.search_address(self.receiver_name_input.text()))


        # Excel Data Placeholder
        self.excel_data = pd.DataFrame()

    def delete_entry(self):
        self.cursor.execute("DELETE FROM orders WHERE id=?", (self.selected_order_id,))
        self.conn.commit()
        self.load_data()
        self.cancel_edit()

    def confirm_delete(self):
        # 경고 대화상자 띄우기
        reply = QMessageBox.question(
            self, '초기화 확인', '정말로 모든 데이터를 삭제하시겠습니까?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, 
            QMessageBox.StandardButton.No  # 기본적으로 No로 설정
        )

        if reply == QMessageBox.StandardButton.Yes:
            self.delete_entry()  # 초기화 실행

    def cancel_edit(self):
        self.sender_name_input.setText("행복한감귤농장")
        self.sender_phone_input.setText("010-5772-4798")
        self.sender_address_input.setText("제주도 서귀포시 남원읍 태위로 603번길 18-5")
        self.receiver_name_input.setText("")
        self.receiver_phone_input.setText("")
        self.receiver_address_input.setText("")
        self.quantity_input.setText("")
        self.add_button.show()
        self.edit_button.hide()
        self.cancel_button.hide()
        self.delete_button.hide()

    def edit_entry(self):
        row_data = [
            self.sender_name_input.text(), self.sender_phone_input.text(), self.sender_address_input.text(),
            self.receiver_name_input.text(), self.receiver_phone_input.text(), self.receiver_address_input.text(),
            self.item_name_input.currentText(), self.quantity_input.text(), self.selected_order_id
        ]
        self.cursor.execute('''
            UPDATE orders SET sender_name=?, sender_phone=?, sender_address=?, 
            receiver_name=?, receiver_phone=?, receiver_address=?, item_name=?, quantity=? 
            WHERE id=?
        ''', row_data)
        self.conn.commit()
        self.load_data()
        self.cancel_edit()

    def table_item_clicked(self, row, column):
        self.selected_order_id = self.table_widget.item(row, 0).text()
        self.sender_name_input.setText(self.table_widget.item(row, 1).text())
        self.sender_phone_input.setText(self.table_widget.item(row, 2).text())
        self.sender_address_input.setText(self.table_widget.item(row, 3).text())
        self.receiver_name_input.setText(self.table_widget.item(row, 5).text())
        self.receiver_phone_input.setText(self.table_widget.item(row, 6).text())
        self.receiver_address_input.setText(self.table_widget.item(row, 7).text())
        self.item_name_input.setCurrentText(self.table_widget.item(row, 9).text())
        self.quantity_input.setText(self.table_widget.item(row, 10).text())
        self.add_button.hide()
        self.edit_button.show()
        self.cancel_button.show()
        self.delete_button.show()

    def confirm_reset(self):
        # 경고 대화상자 띄우기
        reply = QMessageBox.question(
            self, '초기화 확인', '정말로 모든 데이터를 초기화하시겠습니까?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, 
            QMessageBox.StandardButton.No  # 기본적으로 No로 설정
        )

        if reply == QMessageBox.StandardButton.Yes:
            self.reset_all()  # 초기화 실행

    def init_database(self):
        self.conn = sqlite3.connect("addresses.db")
        self.cursor = self.conn.cursor()

        # Orders Table
        self.cursor.execute('''
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sender_name TEXT,
            sender_phone TEXT,
            sender_address TEXT,
            receiver_name TEXT,
            receiver_phone TEXT,
            receiver_address TEXT,
            item_name TEXT,
            quantity INTEGER
        )
        ''')

        # Name-Address Table
        self.cursor.execute('''
        CREATE TABLE IF NOT EXISTS name_address (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            phone TEXT,
            address TEXT
        )
        ''')
        self.conn.commit()

    def load_data(self):
        # 기존 테이블 비우기
        self.table_widget.setRowCount(0)  # 모든 행을 제거
        self.table_widget.clearContents()  # 셀의 내용도 제거

        # 데이터베이스에서 새로운 데이터 로드
        self.cursor.execute("SELECT id, sender_name, sender_phone, sender_address, '', receiver_name, receiver_phone, receiver_address, '', item_name, quantity, '', '', '' FROM orders")
        rows = self.cursor.fetchall()

        # 테이블에 새 데이터 추가
        for row in rows:
            row_position = self.table_widget.rowCount()
            self.table_widget.insertRow(row_position)
            for column, value in enumerate(row):
                self.table_widget.setItem(row_position, column, QTableWidgetItem(str(value)))


    def reset_all(self):
        # 기존 테이블 비우기
        self.table_widget.setRowCount(0)  # 모든 행을 제거
        self.table_widget.clearContents()  # 셀의 내용도 제거
        
        # 데이터베이스에서 모든 데이터 삭제
        self.cursor.execute("DELETE FROM orders")
        self.conn.commit()  # 데이터베이스 변경 사항 저장

    def export_to_excel(self):
        # QTableWidget에서 데이터 가져오기
        rows = self.table_widget.rowCount()
        columns = self.table_widget.columnCount()
        
        data = []
        for row in range(rows):
            row_data = []
            for column in range(1, columns):
                item = self.table_widget.item(row, column)
                row_data.append(item.text() if item is not None else '')
            data.append(row_data)
        
        # pandas DataFrame으로 변환
        df = pd.DataFrame(data, columns=[self.table_widget.horizontalHeaderItem(i).text() for i in range(1, columns)])

        # 엑셀 파일로 저장
        file_path, _ = QFileDialog.getSaveFileName(self, "엑셀 파일 저장", "", "Excel Files (*.xlsx *.xls)")ㄴ
        if not file_path:
            return

        try:
            # pandas의 ExcelWriter로 엑셀 파일 저장
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')

                # 엑셀 파일 열기
                workbook = writer.book
                sheet = workbook['Sheet1']

                # 1열부터 N열까지 가운데 정렬 (3번째 열과 7번째 열 제외)
                for col in range(1, columns + 1):
                    # 3번째 열과 7번째 열은 건너뛰기
                    if col != 3 and col != 7:
                        for row in range(2, rows + 2):  # 첫 번째 행은 헤더이므로 두 번째 행부터 시작
                            sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')

                # 열 너비 자동 조정
                for col in range(1, columns + 1):
                    max_length = 0
                    # 각 열에서 가장 긴 데이터의 길이를 찾아 열 너비를 자동으로 설정
                    for row in range(1, rows + 2):  # 헤더 포함
                        cell_value = str(sheet.cell(row=row, column=col).value)
                        max_length = max(max_length, len(cell_value))
                    adjusted_width = (max_length + 2)  # 여유 공간 추가
                    sheet.column_dimensions[get_column_letter(col)].width = adjusted_width

            QMessageBox.information(self, "Success", "엑셀 파일로 저장되었습니다.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"엑셀 파일 저장 중 오류가 발생했습니다: {e}")



    def add_entry(self):
        row_data = [
            self.sender_name_input.text(), self.sender_phone_input.text(), self.sender_address_input.text(),
            self.receiver_name_input.text(), self.receiver_phone_input.text(), self.receiver_address_input.text(),
            self.item_name_input.currentText(), self.quantity_input.text()
        ]

        # Insert into Orders Table
        self.cursor.execute('''
        INSERT INTO orders (sender_name, sender_phone, sender_address, receiver_name, receiver_phone, receiver_address, item_name, quantity)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', row_data)
        name = self.receiver_name_input.text()
        phone = self.receiver_phone_input.text()
        address = self.receiver_address_input.text()

        if name and phone and address:  # 모든 필드가 존재할 경우만 삽입
            # 중복 확인
            self.cursor.execute('''
            SELECT COUNT(*) FROM name_address
            WHERE name = ? AND phone = ? AND address = ?
            ''', (name, phone, address))
            if self.cursor.fetchone()[0] == 0:  # 데이터가 없을 경우 삽입
                cursor.execute('''
                INSERT INTO name_address (name, phone, address)
                VALUES (?, ?, ?)
                ''', (name, phone, address))
        self.conn.commit()
        self.load_data()
        self.cancel_edit()

    def search_address(self, name):
        self.address_list_widget.clear()
        self.cursor.execute('SELECT phone, address FROM name_address WHERE name LIKE ?', (f"{name}",))
        results = self.cursor.fetchall()
        for phone, address in results:
            self.address_list_widget.addItem(f"{phone} - {address}")

    def fill_address_and_phone_field(self, item):
        text = item.text()
        phone, address = text.split(' - ', 1)
        self.receiver_phone_input.setText(phone)
        self.receiver_address_input.setText(address)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelAddressTool()
    window.show()
    sys.exit(app.exec())